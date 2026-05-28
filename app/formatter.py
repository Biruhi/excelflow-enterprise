# =========================
# FILE: app/formatter.py
# =========================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def format_excel(file_path):

    wb = load_workbook(file_path)

    ws = wb.active

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    # Header styling
    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

    # Auto column widths
    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except:
                pass

        adjusted_width = max_length + 3

        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)